"""
Export service for generating PDF and Excel reports
Supports exporting: borrowing history, member reports, analytics
"""

import io
import csv
from datetime import datetime
from typing import List, Dict, Any
from uuid import UUID

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportService:
    """Service for exporting library data to PDF and Excel"""

    @staticmethod
    def export_borrowing_history_csv(
        borrowings: List[Dict[str, Any]],
        member_name: str = "Member"
    ) -> io.BytesIO:
        """Export borrowing history as CSV"""
        output = io.BytesIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([
            "Book Title",
            "Author",
            "ISBN",
            "Borrowed Date",
            "Due Date",
            "Returned Date",
            "Status",
            "Days Borrowed"
        ])

        # Write data
        for b in borrowings:
            status = "Returned" if b.get("returned_date") else "Active"
            writer.writerow([
                b.get("title", ""),
                b.get("author", ""),
                b.get("isbn", ""),
                b.get("borrowed_date", ""),
                b.get("due_date", ""),
                b.get("returned_date", "") or "N/A",
                status,
                b.get("days_borrowed", "")
            ])

        output.seek(0)
        return output

    @staticmethod
    def export_members_report_csv(
        members: List[Dict[str, Any]]
    ) -> io.BytesIO:
        """Export member list as CSV"""
        output = io.BytesIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([
            "Name",
            "Email",
            "Joined Date",
            "Total Borrowed",
            "Currently Borrowed",
            "Fine Status",
            "Account Status"
        ])

        # Write data
        for m in members:
            writer.writerow([
                m.get("full_name", ""),
                m.get("email", ""),
                m.get("created_at", ""),
                m.get("total_borrowed", 0),
                m.get("currently_borrowed", 0),
                "Unpaid" if m.get("has_unpaid_fines") else "Clear",
                "Suspended" if m.get("is_suspended") else "Active"
            ])

        output.seek(0)
        return output

    @staticmethod
    def export_analytics_csv(
        data: Dict[str, Any]
    ) -> io.BytesIO:
        """Export analytics data as CSV"""
        output = io.BytesIO()
        writer = csv.writer(output)

        # Summary stats
        writer.writerow(["ShelfWise Library - Analytics Report"])
        writer.writerow([datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow([])

        # Summary section
        writer.writerow(["Summary Statistics"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Books", data.get("total_books", 0)])
        writer.writerow(["Total Members", data.get("total_members", 0)])
        writer.writerow(["Active Borrowings", data.get("active_borrowings", 0)])
        writer.writerow(["Borrowed This Month", data.get("borrowed_this_month", 0)])
        writer.writerow([])

        # Top books section
        writer.writerow(["Top Books"])
        writer.writerow(["Title", "Author", "Borrowings"])
        for book in data.get("top_books", []):
            writer.writerow([
                book.get("title", ""),
                book.get("author", ""),
                book.get("count", 0)
            ])
        writer.writerow([])

        # Genre distribution
        writer.writerow(["Genre Distribution"])
        writer.writerow(["Genre", "Count"])
        for genre in data.get("genre_distribution", []):
            writer.writerow([
                genre.get("name", ""),
                genre.get("count", 0)
            ])

        output.seek(0)
        return output

    @staticmethod
    def export_borrowing_history_excel(
        borrowings: List[Dict[str, Any]],
        member_name: str = "Member"
    ) -> io.BytesIO:
        """Export borrowing history as Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed. Install with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Borrowing History"

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Write title
        ws["A1"] = f"Borrowing History - {member_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")

        # Write headers
        headers = ["Book Title", "Author", "ISBN", "Borrowed", "Due Date", "Returned"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Write data
        for row, b in enumerate(borrowings, 4):
            ws.cell(row=row, column=1).value = b.get("title", "")
            ws.cell(row=row, column=2).value = b.get("author", "")
            ws.cell(row=row, column=3).value = b.get("isbn", "")
            ws.cell(row=row, column=4).value = b.get("borrowed_date", "")
            ws.cell(row=row, column=5).value = b.get("due_date", "")
            ws.cell(row=row, column=6).value = b.get("returned_date", "N/A")

            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border

        # Auto-fit columns
        ws.auto_filter.ref = f"A3:F{len(borrowings) + 3}"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_members_report_excel(
        members: List[Dict[str, Any]]
    ) -> io.BytesIO:
        """Export member report as Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "Members"

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 15

        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Title
        ws["A1"] = "ShelfWise Library - Member Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:G1")

        # Headers
        headers = ["Name", "Email", "Joined", "Total Borrowed", "Currently Borrowed", "Fine Status", "Account Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        # Data
        for row, m in enumerate(members, 4):
            ws.cell(row=row, column=1).value = m.get("full_name", "")
            ws.cell(row=row, column=2).value = m.get("email", "")
            ws.cell(row=row, column=3).value = m.get("created_at", "")
            ws.cell(row=row, column=4).value = m.get("total_borrowed", 0)
            ws.cell(row=row, column=5).value = m.get("currently_borrowed", 0)
            ws.cell(row=row, column=6).value = "Unpaid" if m.get("has_unpaid_fines") else "Clear"
            ws.cell(row=row, column=7).value = "Suspended" if m.get("is_suspended") else "Active"

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = border

        ws.auto_filter.ref = f"A3:G{len(members) + 3}"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_analytics_excel(
        data: Dict[str, Any]
    ) -> io.BytesIO:
        """Export analytics as Excel with multiple sheets"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed")

        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = "ShelfWise Library Analytics"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row = 4
        ws_summary.cell(row=row, column=1).value = "Metric"
        ws_summary.cell(row=row, column=2).value = "Value"
        row += 1

        metrics = [
            ("Total Books", data.get("total_books", 0)),
            ("Total Members", data.get("total_members", 0)),
            ("Active Borrowings", data.get("active_borrowings", 0)),
            ("Borrowed This Month", data.get("borrowed_this_month", 0)),
        ]

        for metric, value in metrics:
            ws_summary.cell(row=row, column=1).value = metric
            ws_summary.cell(row=row, column=2).value = value
            row += 1

        # Top Books Sheet
        ws_books = wb.create_sheet("Top Books")
        ws_books["A1"] = "Top Books"
        ws_books["A1"].font = Font(bold=True, size=12)
        ws_books.cell(row=3, column=1).value = "Title"
        ws_books.cell(row=3, column=2).value = "Author"
        ws_books.cell(row=3, column=3).value = "Borrowings"

        for i, book in enumerate(data.get("top_books", []), 4):
            ws_books.cell(row=i, column=1).value = book.get("title", "")
            ws_books.cell(row=i, column=2).value = book.get("author", "")
            ws_books.cell(row=i, column=3).value = book.get("count", 0)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
